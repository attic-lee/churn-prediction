import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
df = pd.read_csv(BASE_DIR / 'outputs' / 'customers_scored.csv')

# ── COLOURS ────────────────────────────────────────────────────────────────
NAVY     = "1B2A4A"
ACCENT   = "2563EB"
WHITE    = "FFFFFF"
GREEN    = "166534"
GREEN_BG = "DCFCE7"
RED      = "991B1B"
RED_BG   = "FEE2E2"
AMBER    = "92400E"
AMBER_BG = "FEF3C7"
GRAY     = "F8FAFC"
MID      = "374151"

def hdr(ws, row, col, text, span=1, color=NAVY):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font      = Font(name="Arial", size=9, bold=True, color=WHITE)
    cell.fill      = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = thin()
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 28
    return cell

def cell(ws, row, col, value, fmt=None, bold=False,
         bg=None, color=MID, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", size=9, bold=bold, color=color)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = thin()
    if fmt: c.number_format = fmt
    if bg:  c.fill = PatternFill("solid", fgColor=bg)
    return c

def thin():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def section(ws, row, col, text, span):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill      = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.border    = thin()
    ws.merge_cells(start_row=row, start_column=col,
                   end_row=row, end_column=col+span-1)
    ws.row_dimensions[row].height = 26

wb = Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — CHURN SUMMARY PIVOT
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Churn_Summary")
ws1.sheet_properties.tabColor = NAVY

section(ws1, 1, 1, "  Churn Summary — by Plan × Contract Type", 6)

# Build pivot
pivot = df.groupby(['plan', 'contract_type']).agg(
    total=('churned_1yes', 'count'),
    churned=('churned_1yes', 'sum'),
    avg_nps=('nps_score_010', 'mean'),
    avg_fee=('monthly_fee_', 'mean')
).reset_index()
pivot['churn_rate'] = pivot['churned'] / pivot['total']

headers = ["Plan", "Contract Type", "Total Customers",
           "Churned", "Churn Rate", "Avg NPS", "Avg Fee (£)"]
for j, h in enumerate(headers, 1):
    hdr(ws1, 2, j, h)

for i, row in pivot.iterrows():
    r = i + 3
    churn_rate = row['churn_rate']
    bg = RED_BG if churn_rate > 0.35 else AMBER_BG if churn_rate > 0.25 else GREEN_BG
    fg = RED    if churn_rate > 0.35 else AMBER    if churn_rate > 0.25 else GREEN

    cell(ws1, r, 1, row['plan'], bold=True)
    cell(ws1, r, 2, row['contract_type'])
    cell(ws1, r, 3, int(row['total']))
    cell(ws1, r, 4, int(row['churned']))
    c = cell(ws1, r, 5, churn_rate, fmt="0.0%", bold=True, bg=bg, color=fg)
    cell(ws1, r, 6, round(row['avg_nps'], 1))
    cell(ws1, r, 7, round(row['avg_fee'], 2), fmt="£#,##0.00")

col_widths = [14, 18, 16, 10, 12, 10, 12]
for j, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(j)].width = w

# KPI summary below pivot
last_row = len(pivot) + 3
ws1.row_dimensions[last_row + 1].height = 10
section(ws1, last_row + 2, 1, "  Overall KPIs", 6)

kpis = [
    ("Total Customers",       len(df)),
    ("Total Churned",         int(df['churned_1yes'].sum())),
    ("Overall Churn Rate",    df['churned_1yes'].mean()),
    ("Monthly Revenue at Risk", df[df['predicted_churn_prob']>0.5]['monthly_fee_'].sum()),
    ("Avg NPS Score",         round(df['nps_score_010'].mean(), 1)),
]
hdr(ws1, last_row + 3, 1, "Metric", span=3)
hdr(ws1, last_row + 3, 4, "Value",  span=3)

fmts = [None, None, "0.0%", "£#,##0.00", None]
for idx, (label, value) in enumerate(kpis):
    r = last_row + 4 + idx
    cell(ws1, r, 1, label, bold=True, align="left", bg=GRAY)
    ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell(ws1, r, 4, value, fmt=fmts[idx], bold=True, bg=GRAY)
    ws1.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)

ws1.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — HIGH RISK SEGMENTS
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("High_Risk_Segments")
ws2.sheet_properties.tabColor = "DC2626"

section(ws2, 1, 1, "  High-Risk Customer List — Predicted Churn Probability > 50%", 9)

high_risk = (df[df['predicted_churn_prob'] > 0.5]
             .sort_values('monthly_fee_', ascending=False)
             .head(100))

# Add recommended intervention
def intervention(row):
    if row['complaints_raised'] >= 2:
        return "Priority: Retention call — address complaint history"
    elif row['nps_score_010'] <= 3:
        return "NPS recovery call — offer service review"
    elif row['contract_type'] == 'Month-to-Month':
        return "Contract upgrade offer — 1-year discount"
    elif row['monthly_fee_'] <= 9.99:
        return "Upgrade offer — Basic to Standard"
    else:
        return "Proactive check-in call"

high_risk = high_risk.copy()
high_risk['intervention'] = high_risk.apply(intervention, axis=1)

headers2 = ["Customer ID", "Plan", "Contract", "Region",
            "Tenure (mo)", "NPS", "Complaints",
            "Monthly Fee (£)", "Churn Probability", "Recommended Intervention"]
for j, h in enumerate(headers2, 1):
    hdr(ws2, 2, j, h)

for i, (_, row) in enumerate(high_risk.iterrows()):
    r = i + 3
    prob = row['predicted_churn_prob']
    bg = RED_BG if prob > 0.75 else AMBER_BG

    cell(ws2, r, 1,  row['customer_id'],  bold=True)
    cell(ws2, r, 2,  row['plan'])
    cell(ws2, r, 3,  row['contract_type'])
    cell(ws2, r, 4,  row['region'])
    cell(ws2, r, 5,  int(row['tenure_months']))
    cell(ws2, r, 6,  int(row['nps_score_010']))
    cell(ws2, r, 7,  int(row['complaints_raised']))
    cell(ws2, r, 8,  row['monthly_fee_'],  fmt="£#,##0.00")
    cell(ws2, r, 9,  prob, fmt="0.0%", bold=True,
         bg=bg, color=RED if prob > 0.75 else AMBER)
    cell(ws2, r, 10, row['intervention'], align="left")

col_widths2 = [12, 10, 16, 14, 12, 6, 10, 14, 18, 40]
for j, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(j)].width = w

ws2.freeze_panes = "A3"

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — MODEL PERFORMANCE
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Model_Performance")
ws3.sheet_properties.tabColor = "2563EB"

section(ws3, 1, 1, "  Model Performance Summary", 4)

hdr(ws3, 2, 1, "Metric")
hdr(ws3, 2, 2, "Logistic Regression")
hdr(ws3, 2, 3, "XGBoost")
hdr(ws3, 2, 4, "Improvement")

metrics = [
    ("AUC-ROC",    0.7531, 0.7837, "=C3-B3"),
    ("Accuracy",   0.76,   0.74,   "=C4-B4"),
    ("Precision",  0.68,   0.56,   "=C5-B5"),
    ("Recall",     0.23,   0.32,   "=C6-B6"),
    ("F1-Score",   0.35,   0.41,   "=C7-B7"),
]
for idx, (name, lr, xgb, diff) in enumerate(metrics):
    r = idx + 3
    cell(ws3, r, 1, name, bold=True, align="left", bg=GRAY)
    cell(ws3, r, 2, lr,  fmt="0.000")
    cell(ws3, r, 3, xgb, fmt="0.000",
         bg=GREEN_BG if xgb > lr else RED_BG,
         color=GREEN if xgb > lr else RED)
    cell(ws3, r, 4, diff, fmt="+0.000;-0.000")

ws3.row_dimensions[8].height = 10
section(ws3, 9, 1, "  Top 5 Churn Predictors (XGBoost Feature Importance)", 4)
hdr(ws3, 10, 1, "Rank")
hdr(ws3, 10, 2, "Feature", span=2)
hdr(ws3, 10, 4, "Business Interpretation")

predictors = [
    (1, "high_risk_flag",        "Customer has low NPS AND a complaint — strongest combined signal"),
    (2, "tenure_band",           "How long the customer has been with the company"),
    (3, "payment_failures_12m",  "Financial stress indicator — missed payments predict churn"),
    (4, "contract_type",         "Month-to-Month customers are 2.3× more likely to churn"),
    (5, "complaints_raised",     "Formal complaints are a direct churn precursor"),
]
for idx, (rank, feature, interpretation) in enumerate(predictors):
    r = idx + 11
    cell(ws3, r, 1, rank, bold=True, bg=GRAY)
    c = ws3.cell(row=r, column=2, value=feature)
    c.font = Font(name="Arial", size=9, bold=True, color=ACCENT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = thin()
    ws3.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cell(ws3, r, 4, interpretation, align="left")

col_widths3 = [20, 22, 10, 50]
for j, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(j)].width = w

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — RECOMMENDATIONS
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Recommendations")
ws4.sheet_properties.tabColor = "166534"

def para(ws, row, text, size=10, bold=False, color=MID, indent=0, height=40):
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(name="Arial", size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal="left", vertical="top",
                            wrap_text=True, indent=indent)
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=6)
    ws.row_dimensions[row].height = height
    return c

ws4.column_dimensions['A'].width = 120

section(ws4, 1, 1, "  Churn Analysis — Consulting Recommendations", 6)

para(ws4, 2,  "CLIENT: UK Telecoms  ·  ANALYST: Attic C. Lee  ·  DATE: March 2026",
     size=9, color=MID, height=20)

para(ws4, 3, "EXECUTIVE SUMMARY", size=11, bold=True, color=NAVY, height=24)

para(ws4, 4,
     "Analysis of 1,000 UK telecoms customers identifies three high-impact retention opportunities. "
     "The model flags 204 customers (20.4%) as high churn risk, representing £4,618/month in revenue "
     "at risk. Three targeted interventions — an onboarding sequence, a Detractor recovery programme, "
     "and a cross-sell campaign — are estimated to reduce monthly churn by 15–20% if implemented "
     "within 30 days.", size=10, color=MID, height=60)

para(ws4, 5, "─" * 120, size=8, color="D1D5DB", height=14)

para(ws4, 6, "FINDING 1 — Early Tenure is the Highest-Risk Window",
     size=11, bold=True, color=RED, height=24)
para(ws4, 7,
     "Customers in their first 6 months churn at 60.3% — nearly 3× the rate of customers "
     "who have been with the company for over 3 years (22.3%). This is the single most "
     "actionable finding: the business is losing customers before they have had a chance to "
     "experience the full product value.", size=10, color=MID, height=55)
para(ws4, 8,
     "Recommendation: Implement a structured 90-day onboarding sequence for all new customers. "
     "Key touchpoints: welcome call at day 3, product tips email at day 14, check-in call at "
     "day 30, satisfaction survey at day 60. Expected impact: 10–15% reduction in 0–6 month "
     "churn based on industry benchmarks for structured onboarding programmes.",
     size=10, color=MID, height=60)

para(ws4, 9, "─" * 120, size=8, color="D1D5DB", height=14)

para(ws4, 10, "FINDING 2 — Detractors on Month-to-Month Contracts Require Immediate Action",
     size=11, bold=True, color=AMBER, height=24)
para(ws4, 11,
     "The high-risk segment — customers with NPS ≤ 3, at least one complaint, and a "
     "Month-to-Month contract — churns at 80%. This 25-customer segment puts £689.75 of "
     "monthly recurring revenue at immediate risk. These customers have already signalled "
     "their dissatisfaction through both survey responses and formal complaints; without "
     "intervention, the majority will leave within 90 days.",
     size=10, color=MID, height=65)
para(ws4, 12,
     "Recommendation: Trigger an immediate retention call campaign for this segment, prioritised "
     "by monthly fee value (see High_Risk_Segments tab). Each call should acknowledge the "
     "complaint history, offer a concrete resolution, and present a contract upgrade incentive "
     "(e.g. 20% discount on 1-year contract). Retaining just 5 of these 25 customers recovers "
     "£138/month — a positive ROI on a single day of retention calls.",
     size=10, color=MID, height=65)

para(ws4, 13, "─" * 120, size=8, color="D1D5DB", height=14)

para(ws4, 14, "FINDING 3 — Multi-Product Customers are Significantly Stickier",
     size=11, bold=True, color=GREEN, height=24)
para(ws4, 15,
     "Customers holding 3 or more products churn at materially lower rates than single-product "
     "customers. The revenue_per_product feature ranks in the top 10 predictors, confirming that "
     "product depth — not just plan tier — drives retention. Single-product Basic plan customers "
     "represent the largest at-risk cohort by volume.",
     size=10, color=MID, height=60)
para(ws4, 16,
     "Recommendation: Launch a targeted cross-sell campaign at month 3 for single-product Basic "
     "plan customers. Offer a bundled upgrade (e.g. Broadband + Mobile at a 15% combined "
     "discount). Month 3 is the optimal intervention point — early enough to build habit before "
     "the churn risk peaks, late enough that the customer has experienced the core product. "
     "Expected impact: moving a customer from 1 to 2 products is estimated to reduce their "
     "churn probability by 20–30% based on the model coefficients.",
     size=10, color=MID, height=70)

para(ws4, 17, "─" * 120, size=8, color="D1D5DB", height=14)

para(ws4, 18, "LIMITATIONS AND NEXT STEPS", size=11, bold=True, color=NAVY, height=24)
para(ws4, 19,
     "This analysis is based on a simulated dataset of 1,000 customers. Model recall (32%) "
     "indicates that the current model misses approximately 68% of actual churners — sufficient "
     "for a pilot retention programme but requiring improvement before full production deployment. "
     "Recommended next steps: (1) collect 12+ months of real transaction data to retrain the "
     "model, (2) add behavioural features such as login frequency and feature usage, "
     "(3) implement A/B testing on the recommended interventions to measure true causal impact.",
     size=10, color=MID, height=80)

# ── Save ────────────────────────────────────────────────────────────────────
output_path = BASE_DIR / 'outputs' / 'Churn_Report.xlsx'
wb.save(output_path)
print(f"Report saved to {output_path}")
print(f"High risk customers in report: {len(high_risk)}")